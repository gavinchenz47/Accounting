import streamlit as st
import pandas as pd
import io
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
import csv

# Page config
st.set_page_config(
    page_title="CRA Payroll Automation",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CRA 2026 Official Rates
class CRA2026:
    CPP_RATE = Decimal('0.0595')
    CPP_EXEMPTION = Decimal('3500.00')
    CPP_MAX_CONTRIBUTION = Decimal('4230.45')
    
    EI_RATE = Decimal('0.0163')
    EI_MAX_PREMIUM = Decimal('1123.07')
    EI_EMPLOYER_MULTIPLIER = Decimal('1.4')
    
    FEDERAL_BPA = Decimal('16452.00')
    ONTARIO_BPA = Decimal('12989.00')
    
    FEDERAL_BRACKETS = [
        (Decimal('58523'), Decimal('0.14'), Decimal('0')),
        (Decimal('117045'), Decimal('0.205'), Decimal('5826.03')),
        (Decimal('181440'), Decimal('0.26'), Decimal('18606.05')),
        (Decimal('258482'), Decimal('0.29'), Decimal('42030.46')),
        (Decimal('999999999'), Decimal('0.33'), Decimal('61367.74'))
    ]
    
    ONTARIO_BRACKETS = [
        (Decimal('53359'), Decimal('0.0505'), Decimal('0')),
        (Decimal('106717'), Decimal('0.0915'), Decimal('2187.67')),
        (Decimal('150000'), Decimal('0.1116'), Decimal('7063.32')),
        (Decimal('220000'), Decimal('0.1216'), Decimal('11893.30')),
        (Decimal('999999999'), Decimal('0.1316'), Decimal('20393.30'))
    ]

class PayrollCalculator:
    def __init__(self, pay_periods=12):
        self.P = Decimal(str(pay_periods))
        self.cra = CRA2026()
    
    def _round(self, amount):
        return Decimal(str(amount)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    def calculate_cpp(self, gross, ytd_cpp):
        exemption = self.cra.CPP_EXEMPTION / self.P
        max_remaining = self.cra.CPP_MAX_CONTRIBUTION - Decimal(str(ytd_cpp))
        period_cpp = self.cra.CPP_RATE * (Decimal(str(gross)) - exemption)
        cpp = min(max_remaining, period_cpp)
        return max(Decimal('0'), self._round(cpp))
    
    def calculate_ei(self, gross, ytd_ei):
        max_remaining = self.cra.EI_MAX_PREMIUM - Decimal(str(ytd_ei))
        period_ei = self.cra.EI_RATE * Decimal(str(gross))
        ei = min(max_remaining, period_ei)
        return max(Decimal('0'), self._round(ei))
    
    def calculate_taxes(self, gross, cpp, ei):
        annual_income = Decimal(str(gross)) * self.P
        
        for bracket_max, rate, constant in self.cra.FEDERAL_BRACKETS:
            if annual_income <= bracket_max:
                R, K = rate, constant
                break
        
        K1 = Decimal('0.14') * self.cra.FEDERAL_BPA
        cpp_credit = self.P * cpp * (Decimal('0.0495') / Decimal('0.0595'))
        ei_credit = self.P * ei
        K2 = Decimal('0.14') * (cpp_credit + ei_credit)
        K4 = min(Decimal('0.14') * annual_income, Decimal('0.14') * Decimal('1501.00'))
        
        T3 = max(Decimal('0'), (R * annual_income) - K - K1 - K2 - K4)
        federal_tax = self._round(T3 / self.P)
        
        for bracket_max, rate, constant in self.cra.ONTARIO_BRACKETS:
            if annual_income <= bracket_max:
                V, KP = rate, constant
                break
        
        K1P = Decimal('0.0505') * self.cra.ONTARIO_BPA
        K2P = Decimal('0.0505') * (cpp_credit + ei_credit)
        
        T4 = max(Decimal('0'), (V * annual_income) - KP - K1P - K2P)
        ohp = self._calculate_ohp(annual_income)
        surtax = self._calculate_surtax(T4)
        
        provincial_tax = self._round((T4 + surtax + ohp) / self.P)
        
        return federal_tax, provincial_tax
    
    def _calculate_ohp(self, annual_income):
        A = annual_income
        if A <= Decimal('20000'):
            return Decimal('0')
        elif A <= Decimal('36000'):
            return min(Decimal('300'), Decimal('0.06') * (A - Decimal('20000')))
        elif A <= Decimal('48000'):
            return min(Decimal('450'), Decimal('300') + Decimal('0.06') * (A - Decimal('36000')))
        elif A <= Decimal('72000'):
            return min(Decimal('600'), Decimal('450') + Decimal('0.25') * (A - Decimal('48000')))
        elif A <= Decimal('200000'):
            return min(Decimal('750'), Decimal('600') + Decimal('0.25') * (A - Decimal('72000')))
        else:
            return min(Decimal('900'), Decimal('750') + Decimal('0.25') * (A - Decimal('200000')))
    
    def _calculate_surtax(self, T4):
        if T4 <= Decimal('5818'):
            return Decimal('0')
        elif T4 <= Decimal('7446'):
            return Decimal('0.20') * (T4 - Decimal('5818'))
        else:
            return Decimal('0.20') * (T4 - Decimal('5818')) + Decimal('0.36') * (T4 - Decimal('7446'))
    
    def calculate_payroll(self, employee):
        gross = Decimal(str(employee['gross_pay']))
        ytd_cpp = Decimal(str(employee.get('ytd_cpp', 0)))
        ytd_ei = Decimal(str(employee.get('ytd_ei', 0)))
        
        cpp = self.calculate_cpp(gross, ytd_cpp)
        ei = self.calculate_ei(gross, ytd_ei)
        federal_tax, provincial_tax = self.calculate_taxes(gross, cpp, ei)
        
        total_deductions = cpp + ei + federal_tax + provincial_tax
        net_amount = gross - total_deductions
        
        employer_cpp = cpp
        employer_ei = self._round(ei * self.cra.EI_EMPLOYER_MULTIPLIER)
        employer_total = employer_cpp + employer_ei
        remittance_total = total_deductions + employer_total
        
        return {
            'name': employee['name'],
            'gross_pay': self._round(gross),
            'cpp': cpp,
            'ei': ei,
            'federal_tax': federal_tax,
            'provincial_tax': provincial_tax,
            'total_deductions': total_deductions,
            'net_amount': net_amount,
            'employer_cpp': employer_cpp,
            'employer_ei': employer_ei,
            'employer_total': employer_total,
            'remittance_total': remittance_total,
            'ytd_cpp_new': self._round(ytd_cpp + cpp),
            'ytd_ei_new': self._round(ytd_ei + ei)
        }

def create_excel_output(payroll_data, pay_period):
    """Create Excel file in memory"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Summary"
    
    # Title
    ws['A1'] = 'PAYROLL DEDUCTIONS SUMMARY'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws.merge_cells('A1:L1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = f'Pay Period: {pay_period}'
    ws['A2'].font = Font(italic=True, size=11)
    ws.merge_cells('A2:L2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Employee', 'Gross Pay', 'CPP (Employee)', 'EI (Employee)', 
               'Federal Tax', 'Provincial Tax', 'Total Deductions', 'Net Amount',
               'CPP (Employer)', 'EI (Employer)', 'Employer Total', 'Remit to CRA']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Data
    row = 5
    for emp in payroll_data:
        ws.cell(row=row, column=1, value=emp['name'])
        ws.cell(row=row, column=2, value=float(emp['gross_pay'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=3, value=float(emp['cpp'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=4, value=float(emp['ei'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=5, value=float(emp['federal_tax'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=6, value=float(emp['provincial_tax'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=7, value=float(emp['total_deductions'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=8, value=float(emp['net_amount'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=9, value=float(emp['employer_cpp'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=10, value=float(emp['employer_ei'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=11, value=float(emp['employer_total'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=12, value=float(emp['remittance_total'])).number_format = '$#,##0.00'
        
        ws.cell(row=row, column=8).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        ws.cell(row=row, column=8).font = Font(bold=True, color='375623')
        row += 1
    
    # Total row
    total_row = row
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True, size=11)
    for col in range(2, 13):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f'=SUM({chr(64+col)}5:{chr(64+col)}{total_row-1})'
        cell.font = Font(bold=True, size=11)
        cell.number_format = '$#,##0.00'
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    for col in 'BCDEFGHIJKL':
        ws.column_dimensions[col].width = 14
    
    # T4 Sheet
    ws_t4 = wb.create_sheet(title="T4 Preparation")
    ws_t4['A1'] = 'T4 SLIP PREPARATION DATA'
    ws_t4['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws_t4['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws_t4.merge_cells('A1:G1')
    ws_t4['A1'].alignment = Alignment(horizontal='center')
    
    t4_headers = ['Employee Name', 'Box 14\nEmployment Income', 'Box 16\nCPP', 
                  'Box 18\nEI', 'Box 22\nIncome Tax', 'YTD CPP', 'YTD EI']
    
    for col, header in enumerate(t4_headers, 1):
        cell = ws_t4.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    row = 4
    for emp in payroll_data:
        ws_t4.cell(row=row, column=1, value=emp['name'])
        ws_t4.cell(row=row, column=2, value=float(emp['gross_pay'])).number_format = '$#,##0.00'
        ws_t4.cell(row=row, column=3, value=float(emp['cpp'])).number_format = '$#,##0.00'
        ws_t4.cell(row=row, column=4, value=float(emp['ei'])).number_format = '$#,##0.00'
        ws_t4.cell(row=row, column=5, value=float(emp['federal_tax'] + emp['provincial_tax'])).number_format = '$#,##0.00'
        ws_t4.cell(row=row, column=6, value=float(emp['ytd_cpp_new'])).number_format = '$#,##0.00'
        ws_t4.cell(row=row, column=7, value=float(emp['ytd_ei_new'])).number_format = '$#,##0.00'
        row += 1
    
    ws_t4.column_dimensions['A'].width = 20
    for col in 'BCDEFG':
        ws_t4.column_dimensions[col].width = 16
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Main App
def main():
    # Sidebar
    with st.sidebar:
        st.title("💼 CRA Payroll")
        st.markdown("---")
        st.markdown("### About")
        st.info("Automates CRA payroll calculations. Upload a CSV or enter employee data manually to get T4-ready Excel reports.")
        
        st.markdown("### 2026 CRA Rates")
        st.text(f"CPP: 5.95%")
        st.text(f"EI: 1.63%")
        st.text(f"Federal Tax: 14% (base)")
        
        st.markdown("---")
        st.markdown("### Need Help?")
        st.markdown("📄 Upload a CSV with columns:")
        st.markdown("`Employee Name, Gross Pay, YTD CPP, YTD EI, Pay Periods`")
    
    # Header
    st.title("🇨🇦 CRA Payroll Automation Tool")
    st.markdown("**Automated payroll calculations matching the CRA website - Get T4-ready Excel in seconds**")
    st.markdown("---")
    
    # Pay period (shared across both tabs)
    pay_period = st.text_input(
        "Pay Period",
        value=datetime.now().strftime("%B %Y"),
        help="e.g., January 2026"
    )

    # Input tabs
    tab_manual, tab_csv = st.tabs(["✏️ Manual Entry", "📤 Upload CSV"])

    employees = None

    with tab_csv:
        with st.expander("📋 CSV Format", expanded=False):
            st.markdown("""
            CSV must have columns: `Employee Name, Gross Pay, YTD CPP, YTD EI, Pay Periods`

            **YTD (Year-to-Date) Tips:**
            - First month of year: YTD CPP = 0, YTD EI = 0
            - Subsequent months: Use YTD values from previous month's output
            """)

        uploaded_file = st.file_uploader(
            "Choose CSV file",
            type=['csv'],
            help="CSV must have columns: Employee Name, Gross Pay, YTD CPP, YTD EI, Pay Periods"
        )

        if uploaded_file and st.button("🧮 Calculate Payroll", type="primary", use_container_width=True, key="csv_calc"):
            try:
                df = pd.read_csv(uploaded_file)
                required_cols = ['Employee Name', 'Gross Pay', 'YTD CPP', 'YTD EI', 'Pay Periods']
                missing = [col for col in required_cols if col not in df.columns]
                if missing:
                    st.error(f"❌ Missing columns: {', '.join(missing)}")
                else:
                    employees = []
                    has_errors = False
                    for idx, row in df.iterrows():
                        ytd_cpp = float(row['YTD CPP'])
                        ytd_ei = float(row['YTD EI'])
                        gross = float(str(row['Gross Pay']).replace(',', '').replace('$', ''))
                        name = str(row['Employee Name']).strip()
                        if ytd_cpp > 4230.45:
                            st.error(f"❌ {name}: YTD CPP (${ytd_cpp:,.2f}) exceeds 2026 maximum of $4,230.45")
                            has_errors = True
                        if ytd_ei > 1123.07:
                            st.error(f"❌ {name}: YTD EI (${ytd_ei:,.2f}) exceeds 2026 maximum of $1,123.07")
                            has_errors = True
                        if gross <= 0:
                            st.error(f"❌ {name}: Gross Pay must be greater than 0.")
                            has_errors = True
                        employees.append({
                            'name': name,
                            'gross_pay': gross,
                            'ytd_cpp': ytd_cpp,
                            'ytd_ei': ytd_ei,
                            'pay_periods': int(row['Pay Periods'])
                        })
                    if has_errors:
                        employees = None
            except Exception as e:
                st.error(f"❌ Error reading file: {str(e)}")
                st.info("Please check your CSV format matches the template.")

    with tab_manual:
        st.markdown("Enter employee data directly. Use **+ Add Employee** to add more rows (up to 10).")
        st.caption("YTD = Year-to-Date. First pay period of the year? Leave YTD fields at 0.")

        CPP_MAX = 4230.45
        EI_MAX = 1123.07

        if 'num_employees' not in st.session_state:
            st.session_state.num_employees = 1

        col_add, col_remove = st.columns(2)
        with col_add:
            if st.button("+ Add Employee", use_container_width=True):
                st.session_state.num_employees = min(st.session_state.num_employees + 1, 10)
                st.rerun()
        with col_remove:
            if st.session_state.num_employees > 1:
                if st.button("- Remove Last", use_container_width=True):
                    st.session_state.num_employees -= 1
                    st.rerun()

        pay_period_options = {"Monthly (12)": 12, "Semi-Monthly (24)": 24, "Bi-Weekly (26)": 26, "Weekly (52)": 52}

        manual_data = []
        for i in range(st.session_state.num_employees):
            st.markdown(f"---")
            st.markdown(f"**Employee {i + 1}**")
            c1, c2 = st.columns(2)
            with c1:
                name = st.text_input("Employee Name", key=f"name_{i}", placeholder="e.g. John Smith")
            with c2:
                pp = st.selectbox("Pay Frequency", options=list(pay_period_options.keys()), key=f"pp_{i}")
            c3, c4, c5 = st.columns(3)
            with c3:
                gross_str = st.text_input("Gross Pay ($)", key=f"gross_{i}", placeholder="e.g. 5000.00",
                                          help="This period's gross salary before deductions")
            with c4:
                ytd_cpp_str = st.text_input(f"YTD CPP (max ${CPP_MAX:,.2f})", key=f"ytd_cpp_{i}", placeholder="0.00",
                                            help="Total CPP already deducted this year from prior pay periods")
            with c5:
                ytd_ei_str = st.text_input(f"YTD EI (max ${EI_MAX:,.2f})", key=f"ytd_ei_{i}", placeholder="0.00",
                                           help="Total EI already deducted this year from prior pay periods")
            manual_data.append({'name': name, 'gross_str': gross_str, 'ytd_cpp_str': ytd_cpp_str, 'ytd_ei_str': ytd_ei_str, 'pp_label': pp})

        st.markdown("---")
        if st.button("🧮 Calculate Payroll", type="primary", use_container_width=True, key="manual_calc"):
            valid = True
            parsed = []
            for i, d in enumerate(manual_data):
                if not d['name'].strip():
                    st.error(f"❌ Employee {i + 1}: Name is required.")
                    valid = False

                # Parse gross pay
                gross_raw = d['gross_str'].strip().replace(',', '').replace('$', '')
                if not gross_raw:
                    st.error(f"❌ Employee {i + 1}: Gross Pay is required.")
                    valid = False
                    gross = 0.0
                else:
                    try:
                        gross = float(gross_raw)
                        if gross <= 0:
                            st.error(f"❌ Employee {i + 1}: Gross Pay must be greater than 0.")
                            valid = False
                    except ValueError:
                        st.error(f"❌ Employee {i + 1}: Gross Pay must be a number.")
                        valid = False
                        gross = 0.0

                # Parse YTD CPP
                cpp_raw = d['ytd_cpp_str'].strip().replace(',', '').replace('$', '')
                if not cpp_raw:
                    ytd_cpp = 0.0
                else:
                    try:
                        ytd_cpp = float(cpp_raw)
                        if ytd_cpp < 0:
                            st.error(f"❌ Employee {i + 1}: YTD CPP cannot be negative.")
                            valid = False
                        elif ytd_cpp > CPP_MAX:
                            st.error(f"❌ Employee {i + 1}: YTD CPP (${ytd_cpp:,.2f}) exceeds maximum of ${CPP_MAX:,.2f}.")
                            valid = False
                    except ValueError:
                        st.error(f"❌ Employee {i + 1}: YTD CPP must be a number.")
                        valid = False
                        ytd_cpp = 0.0

                # Parse YTD EI
                ei_raw = d['ytd_ei_str'].strip().replace(',', '').replace('$', '')
                if not ei_raw:
                    ytd_ei = 0.0
                else:
                    try:
                        ytd_ei = float(ei_raw)
                        if ytd_ei < 0:
                            st.error(f"❌ Employee {i + 1}: YTD EI cannot be negative.")
                            valid = False
                        elif ytd_ei > EI_MAX:
                            st.error(f"❌ Employee {i + 1}: YTD EI (${ytd_ei:,.2f}) exceeds maximum of ${EI_MAX:,.2f}.")
                            valid = False
                    except ValueError:
                        st.error(f"❌ Employee {i + 1}: YTD EI must be a number.")
                        valid = False
                        ytd_ei = 0.0

                parsed.append({'name': d['name'].strip(), 'gross': gross, 'ytd_cpp': ytd_cpp, 'ytd_ei': ytd_ei, 'pp_label': d['pp_label']})

            if valid:
                employees = []
                for d in parsed:
                    employees.append({
                        'name': d['name'],
                        'gross_pay': d['gross'],
                        'ytd_cpp': d['ytd_cpp'],
                        'ytd_ei': d['ytd_ei'],
                        'pay_periods': pay_period_options[d['pp_label']]
                    })

    # Process employees (from either tab)
    if employees:
        try:
            with st.spinner("🔄 Processing payroll calculations..."):
                results = []
                for emp in employees:
                    calculator = PayrollCalculator(pay_periods=emp['pay_periods'])
                    results.append(calculator.calculate_payroll(emp))

            st.success(f"✅ Processed {len(results)} employees successfully!")

            st.markdown("---")
            st.subheader("📊 Payroll Summary")

            total_gross = sum(float(r['gross_pay']) for r in results)
            total_net = sum(float(r['net_amount']) for r in results)
            total_remit = sum(float(r['remittance_total']) for r in results)

            col1, col2, col3 = st.columns(3)
            col1.metric("Total Gross Pay", f"${total_gross:,.2f}")
            col2.metric("Total Net Pay", f"${total_net:,.2f}")
            col3.metric("Total to Remit to CRA", f"${total_remit:,.2f}",
                       help="Includes all deductions + employer CPP + employer EI")

            st.markdown("### Employee Details")
            results_df = pd.DataFrame([
                {
                    'Employee': r['name'],
                    'Gross Pay': f"${float(r['gross_pay']):,.2f}",
                    'CPP': f"${float(r['cpp']):,.2f}",
                    'EI': f"${float(r['ei']):,.2f}",
                    'Federal Tax': f"${float(r['federal_tax']):,.2f}",
                    'Provincial Tax': f"${float(r['provincial_tax']):,.2f}",
                    'Net Amount': f"${float(r['net_amount']):,.2f}",
                    'CRA Remittance': f"${float(r['remittance_total']):,.2f}"
                }
                for r in results
            ])

            st.dataframe(results_df, use_container_width=True, hide_index=True)

            st.markdown("---")
            excel_data = create_excel_output(results, pay_period)
            filename = f"payroll_{pay_period.replace(' ', '_').lower()}.xlsx"

            st.download_button(
                label="📥 Download Complete Excel Report",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

            st.info("💡 Excel file contains: **Payroll Summary** + **T4 Preparation** sheets")

        except Exception as e:
            st.error(f"❌ Error processing payroll: {str(e)}")
    
    # Footer
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: #666; font-size: 0.9em;'>
    <p>CRA-Compliant Calculations | Based on T4127 Payroll Deductions Formulas (122nd Edition - January 2026)</p>
    <p>⚠️ This tool is for estimation purposes. Always verify with a tax professional.</p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
