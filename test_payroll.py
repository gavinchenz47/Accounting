"""Tests for CRA 2026 payroll calculator."""

import pytest
from decimal import Decimal

from payroll import PayrollCalculator, CRA2026, create_excel_output
from provinces import PROVINCES, get_health_premium, get_surtax


# ── CRA2026 constants ──

class TestCRA2026Constants:
    def test_cpp_rate(self):
        assert CRA2026.CPP_RATE == Decimal('0.0595')

    def test_cpp_exemption(self):
        assert CRA2026.CPP_EXEMPTION == Decimal('3500.00')

    def test_cpp_max_contribution(self):
        assert CRA2026.CPP_MAX_CONTRIBUTION == Decimal('4230.45')

    def test_ei_rate(self):
        assert CRA2026.EI_RATE == Decimal('0.0163')

    def test_ei_max_premium(self):
        assert CRA2026.EI_MAX_PREMIUM == Decimal('1123.07')

    def test_ei_employer_multiplier(self):
        assert CRA2026.EI_EMPLOYER_MULTIPLIER == Decimal('1.4')

    def test_federal_bpa(self):
        assert CRA2026.FEDERAL_BPA == Decimal('16452.00')

    def test_ontario_bpa(self):
        assert PROVINCES['ON']['bpa'] == Decimal('12989.00')

    def test_federal_brackets_count(self):
        assert len(CRA2026.FEDERAL_BRACKETS) == 5

    def test_ontario_brackets_count(self):
        assert len(PROVINCES['ON']['brackets']) == 5


# ── CPP calculations ──

class TestCPPCalculation:
    def test_basic_cpp_monthly(self):
        calc = PayrollCalculator(pay_periods=12)
        cpp = calc.calculate_cpp(5000, 0)
        # (5000 - 3500/12) * 0.0595 = (5000 - 291.67) * 0.0595 = 280.15
        assert cpp == Decimal('280.15')

    def test_cpp_with_ytd_near_max(self):
        """CPP should be capped when YTD is near annual maximum."""
        calc = PayrollCalculator(pay_periods=12)
        cpp = calc.calculate_cpp(5000, 4200)
        # max_remaining = 4230.45 - 4200 = 30.45
        # period_cpp = 280.15 (from above)
        # min(30.45, 280.15) = 30.45
        assert cpp == Decimal('30.45')

    def test_cpp_at_max(self):
        """CPP should be 0 when YTD equals annual maximum."""
        calc = PayrollCalculator(pay_periods=12)
        cpp = calc.calculate_cpp(5000, 4230.45)
        assert cpp == Decimal('0.00')

    def test_cpp_over_max(self):
        """CPP should be 0 when YTD exceeds annual maximum."""
        calc = PayrollCalculator(pay_periods=12)
        cpp = calc.calculate_cpp(5000, 4300)
        assert cpp == Decimal('0.00')

    def test_cpp_low_income(self):
        """CPP should be 0 when gross is below per-period exemption."""
        calc = PayrollCalculator(pay_periods=12)
        # exemption per period = 3500/12 = 291.67
        cpp = calc.calculate_cpp(200, 0)
        assert cpp == Decimal('0.00')

    def test_cpp_biweekly(self):
        """CPP should use correct exemption for biweekly pay periods."""
        calc = PayrollCalculator(pay_periods=26)
        cpp = calc.calculate_cpp(2500, 0)
        # exemption per period = 3500/26 = 134.615...
        # (2500 - 134.615...) * 0.0595 = 140.74
        assert cpp == Decimal('140.74')

    def test_cpp_weekly(self):
        calc = PayrollCalculator(pay_periods=52)
        cpp = calc.calculate_cpp(1200, 0)
        # exemption per period = 3500/52 = 67.307692...
        # (1200 - 67.307692...) * 0.0595 = 67.40
        assert cpp == Decimal('67.40')


# ── EI calculations ──

class TestEICalculation:
    def test_basic_ei_monthly(self):
        calc = PayrollCalculator(pay_periods=12)
        ei = calc.calculate_ei(5000, 0)
        # 5000 * 0.0163 = 81.50
        assert ei == Decimal('81.50')

    def test_ei_with_ytd_near_max(self):
        calc = PayrollCalculator(pay_periods=12)
        ei = calc.calculate_ei(5000, 1100)
        # max_remaining = 1123.07 - 1100 = 23.07
        assert ei == Decimal('23.07')

    def test_ei_at_max(self):
        calc = PayrollCalculator(pay_periods=12)
        ei = calc.calculate_ei(5000, 1123.07)
        assert ei == Decimal('0.00')

    def test_ei_over_max(self):
        calc = PayrollCalculator(pay_periods=12)
        ei = calc.calculate_ei(5000, 1200)
        assert ei == Decimal('0.00')

    def test_ei_biweekly(self):
        calc = PayrollCalculator(pay_periods=26)
        ei = calc.calculate_ei(2500, 0)
        # 2500 * 0.0163 = 40.75
        assert ei == Decimal('40.75')


# ── Ontario Health Premium ──

class TestOntarioHealthPremium:
    def test_ohp_below_20000(self):
        assert get_health_premium('ON', Decimal('15000')) == Decimal('0')

    def test_ohp_at_20000(self):
        assert get_health_premium('ON', Decimal('20000')) == Decimal('0')

    def test_ohp_between_20000_and_36000(self):
        assert get_health_premium('ON', Decimal('30000')) == Decimal('300')

    def test_ohp_at_25000(self):
        assert get_health_premium('ON', Decimal('25000')) == Decimal('300')

    def test_ohp_between_36000_and_48000(self):
        assert get_health_premium('ON', Decimal('42000')) == Decimal('450')

    def test_ohp_between_48000_and_72000(self):
        assert get_health_premium('ON', Decimal('60000')) == Decimal('600')

    def test_ohp_between_72000_and_200000(self):
        assert get_health_premium('ON', Decimal('100000')) == Decimal('750')

    def test_ohp_above_200000(self):
        assert get_health_premium('ON', Decimal('250000')) == Decimal('900')

    def test_no_ohp_for_other_provinces(self):
        assert get_health_premium('BC', Decimal('100000')) == Decimal('0')
        assert get_health_premium('AB', Decimal('100000')) == Decimal('0')


# ── Ontario Surtax ──

class TestOntarioSurtax:
    def test_surtax_below_threshold(self):
        assert get_surtax('ON', Decimal('5000')) == Decimal('0')

    def test_surtax_at_5818(self):
        assert get_surtax('ON', Decimal('5818')) == Decimal('0')

    def test_surtax_between_5818_and_7446(self):
        result = get_surtax('ON', Decimal('6500'))
        assert result == Decimal('136.40')

    def test_surtax_above_7446(self):
        result = get_surtax('ON', Decimal('8000'))
        assert result == Decimal('635.84')

    def test_no_surtax_for_alberta(self):
        assert get_surtax('AB', Decimal('50000')) == Decimal('0')


# ── Full payroll calculation ──

class TestFullPayroll:
    def test_monthly_5000_gross(self):
        calc = PayrollCalculator(pay_periods=12)
        result = calc.calculate_payroll({
            'name': 'John Smith',
            'gross_pay': 5000,
            'ytd_cpp': 0,
            'ytd_ei': 0
        })
        assert result['name'] == 'John Smith'
        assert result['gross_pay'] == Decimal('5000.00')
        assert result['cpp'] == Decimal('280.15')
        assert result['ei'] == Decimal('81.50')
        assert result['employer_cpp'] == result['cpp']
        assert result['employer_ei'] == Decimal('114.10')  # 81.50 * 1.4
        assert result['net_amount'] == result['gross_pay'] - result['total_deductions']
        assert result['ytd_cpp_new'] == Decimal('280.15')
        assert result['ytd_ei_new'] == Decimal('81.50')

    def test_monthly_6500_gross(self):
        calc = PayrollCalculator(pay_periods=12)
        result = calc.calculate_payroll({
            'name': 'Jane Doe',
            'gross_pay': 6500,
            'ytd_cpp': 0,
            'ytd_ei': 0
        })
        assert result['gross_pay'] == Decimal('6500.00')
        assert result['cpp'] > Decimal('0')
        assert result['ei'] > Decimal('0')
        assert result['federal_tax'] > Decimal('0')
        assert result['provincial_tax'] > Decimal('0')

    def test_deductions_sum_correctly(self):
        calc = PayrollCalculator(pay_periods=12)
        result = calc.calculate_payroll({
            'name': 'Test',
            'gross_pay': 5000,
            'ytd_cpp': 0,
            'ytd_ei': 0
        })
        expected_deductions = result['cpp'] + result['ei'] + result['federal_tax'] + result['provincial_tax']
        assert result['total_deductions'] == expected_deductions

    def test_net_amount_is_gross_minus_deductions(self):
        calc = PayrollCalculator(pay_periods=12)
        result = calc.calculate_payroll({
            'name': 'Test',
            'gross_pay': 7000,
            'ytd_cpp': 0,
            'ytd_ei': 0
        })
        assert result['net_amount'] == result['gross_pay'] - result['total_deductions']

    def test_employer_ei_is_1_4x(self):
        calc = PayrollCalculator(pay_periods=12)
        result = calc.calculate_payroll({
            'name': 'Test',
            'gross_pay': 5000,
            'ytd_cpp': 0,
            'ytd_ei': 0
        })
        expected_employer_ei = (result['ei'] * Decimal('1.4')).quantize(Decimal('0.01'))
        assert result['employer_ei'] == expected_employer_ei

    def test_employer_cpp_matches_employee(self):
        calc = PayrollCalculator(pay_periods=12)
        result = calc.calculate_payroll({
            'name': 'Test',
            'gross_pay': 5000,
            'ytd_cpp': 0,
            'ytd_ei': 0
        })
        assert result['employer_cpp'] == result['cpp']

    def test_remittance_includes_employer_portions(self):
        calc = PayrollCalculator(pay_periods=12)
        result = calc.calculate_payroll({
            'name': 'Test',
            'gross_pay': 5000,
            'ytd_cpp': 0,
            'ytd_ei': 0
        })
        expected = result['total_deductions'] + result['employer_total']
        assert result['remittance_total'] == expected

    def test_ytd_tracking(self):
        """YTD values should accumulate correctly across pay periods."""
        calc = PayrollCalculator(pay_periods=12)
        # January
        r1 = calc.calculate_payroll({'name': 'Test', 'gross_pay': 5000, 'ytd_cpp': 0, 'ytd_ei': 0})
        # February (using January's new YTD)
        r2 = calc.calculate_payroll({
            'name': 'Test',
            'gross_pay': 5000,
            'ytd_cpp': float(r1['ytd_cpp_new']),
            'ytd_ei': float(r1['ytd_ei_new'])
        })
        assert r2['ytd_cpp_new'] == r1['ytd_cpp_new'] + r2['cpp']
        assert r2['ytd_ei_new'] == r1['ytd_ei_new'] + r2['ei']

    def test_cpp_stops_at_annual_max_over_months(self):
        """CPP deductions should stop once annual max is reached across months."""
        calc = PayrollCalculator(pay_periods=12)
        ytd_cpp = Decimal('0')
        ytd_ei = Decimal('0')
        total_cpp = Decimal('0')
        for _ in range(12):
            result = calc.calculate_payroll({
                'name': 'Test',
                'gross_pay': 8000,
                'ytd_cpp': float(ytd_cpp),
                'ytd_ei': float(ytd_ei)
            })
            total_cpp += result['cpp']
            ytd_cpp = result['ytd_cpp_new']
            ytd_ei = result['ytd_ei_new']
        assert total_cpp == Decimal('4230.45')

    def test_ei_stops_at_annual_max_over_months(self):
        """EI deductions should stop once annual max is reached across months."""
        calc = PayrollCalculator(pay_periods=12)
        ytd_cpp = Decimal('0')
        ytd_ei = Decimal('0')
        total_ei = Decimal('0')
        for _ in range(12):
            result = calc.calculate_payroll({
                'name': 'Test',
                'gross_pay': 8000,
                'ytd_cpp': float(ytd_cpp),
                'ytd_ei': float(ytd_ei)
            })
            total_ei += result['ei']
            ytd_cpp = result['ytd_cpp_new']
            ytd_ei = result['ytd_ei_new']
        assert total_ei == Decimal('1123.07')


# ── Pay period variations ──

class TestPayPeriodVariations:
    def test_biweekly_produces_different_results(self):
        monthly = PayrollCalculator(pay_periods=12)
        biweekly = PayrollCalculator(pay_periods=26)
        r_monthly = monthly.calculate_payroll({'name': 'Test', 'gross_pay': 5000, 'ytd_cpp': 0, 'ytd_ei': 0})
        r_biweekly = biweekly.calculate_payroll({'name': 'Test', 'gross_pay': 5000, 'ytd_cpp': 0, 'ytd_ei': 0})
        # Same gross but different pay periods should produce different tax
        assert r_monthly['federal_tax'] != r_biweekly['federal_tax']

    def test_weekly_pay_period(self):
        calc = PayrollCalculator(pay_periods=52)
        result = calc.calculate_payroll({'name': 'Test', 'gross_pay': 1200, 'ytd_cpp': 0, 'ytd_ei': 0})
        assert result['cpp'] > Decimal('0')
        assert result['ei'] > Decimal('0')
        assert result['net_amount'] > Decimal('0')

    def test_semi_monthly(self):
        calc = PayrollCalculator(pay_periods=24)
        result = calc.calculate_payroll({'name': 'Test', 'gross_pay': 2500, 'ytd_cpp': 0, 'ytd_ei': 0})
        assert result['cpp'] > Decimal('0')
        assert result['net_amount'] > Decimal('0')


# ── Edge cases ──

class TestEdgeCases:
    def test_very_high_income(self):
        """High income should hit top tax brackets."""
        calc = PayrollCalculator(pay_periods=12)
        result = calc.calculate_payroll({'name': 'CEO', 'gross_pay': 50000, 'ytd_cpp': 0, 'ytd_ei': 0})
        assert result['federal_tax'] > Decimal('0')
        assert result['provincial_tax'] > Decimal('0')
        assert result['net_amount'] > Decimal('0')
        assert result['net_amount'] < result['gross_pay']

    def test_low_income_no_tax(self):
        """Very low income should result in zero or minimal tax."""
        calc = PayrollCalculator(pay_periods=12)
        result = calc.calculate_payroll({'name': 'PartTime', 'gross_pay': 500, 'ytd_cpp': 0, 'ytd_ei': 0})
        assert result['federal_tax'] == Decimal('0.00')
        assert result['net_amount'] > Decimal('0')

    def test_minimum_wage_monthly(self):
        """Minimum wage worker should get reasonable deductions."""
        # ~$15.50/hr * 160 hrs = $2480/month
        calc = PayrollCalculator(pay_periods=12)
        result = calc.calculate_payroll({'name': 'MinWage', 'gross_pay': 2480, 'ytd_cpp': 0, 'ytd_ei': 0})
        assert result['net_amount'] > Decimal('0')
        assert result['cpp'] > Decimal('0')
        assert result['ei'] > Decimal('0')

    def test_defaults_for_missing_ytd(self):
        """Missing YTD fields should default to 0."""
        calc = PayrollCalculator(pay_periods=12)
        result = calc.calculate_payroll({'name': 'Test', 'gross_pay': 5000})
        assert result['cpp'] == Decimal('280.15')
        assert result['ei'] == Decimal('81.50')


# ── Excel output ──

class TestExcelOutput:
    def test_creates_valid_excel(self):
        calc = PayrollCalculator(pay_periods=12)
        results = [
            calc.calculate_payroll({'name': 'John Smith', 'gross_pay': 5000, 'ytd_cpp': 0, 'ytd_ei': 0}),
            calc.calculate_payroll({'name': 'Jane Doe', 'gross_pay': 6500, 'ytd_cpp': 0, 'ytd_ei': 0}),
        ]
        excel_data = create_excel_output(results, "January 2026")
        assert excel_data is not None
        assert len(excel_data) > 0

    def test_excel_has_two_sheets(self):
        from openpyxl import load_workbook
        import io
        calc = PayrollCalculator(pay_periods=12)
        results = [
            calc.calculate_payroll({'name': 'Test', 'gross_pay': 5000, 'ytd_cpp': 0, 'ytd_ei': 0}),
        ]
        excel_data = create_excel_output(results, "January 2026")
        wb = load_workbook(io.BytesIO(excel_data))
        assert len(wb.sheetnames) == 2
        assert "Payroll Summary" in wb.sheetnames
        assert "T4 Preparation" in wb.sheetnames

    def test_excel_employee_count_matches(self):
        from openpyxl import load_workbook
        import io
        calc = PayrollCalculator(pay_periods=12)
        results = [
            calc.calculate_payroll({'name': f'Emp {i}', 'gross_pay': 5000, 'ytd_cpp': 0, 'ytd_ei': 0})
            for i in range(3)
        ]
        excel_data = create_excel_output(results, "March 2026")
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb["Payroll Summary"]
        assert ws.cell(row=5, column=1).value == 'Emp 0'
        assert ws.cell(row=7, column=1).value == 'Emp 2'
        assert ws.cell(row=8, column=1).value == 'TOTAL'


# ── Province variations ──

class TestProvinceVariations:
    def test_all_provinces_load(self):
        """All 13 provinces/territories should be available."""
        assert len(PROVINCES) == 13

    def test_different_provinces_different_tax(self):
        """Same income in different provinces should produce different provincial tax."""
        emp = {'name': 'Test', 'gross_pay': 5000, 'ytd_cpp': 0, 'ytd_ei': 0}
        on_calc = PayrollCalculator(pay_periods=12, province='ON')
        ab_calc = PayrollCalculator(pay_periods=12, province='AB')
        bc_calc = PayrollCalculator(pay_periods=12, province='BC')
        on_result = on_calc.calculate_payroll(emp)
        ab_result = ab_calc.calculate_payroll(emp)
        bc_result = bc_calc.calculate_payroll(emp)
        # Provincial taxes should differ
        assert on_result['provincial_tax'] != ab_result['provincial_tax']
        assert on_result['provincial_tax'] != bc_result['provincial_tax']
        # Federal tax should be the same
        assert on_result['federal_tax'] == ab_result['federal_tax']
        assert on_result['federal_tax'] == bc_result['federal_tax']

    def test_alberta_higher_bpa_lower_tax(self):
        """Alberta's high BPA ($21,003) should result in lower tax for low income."""
        emp = {'name': 'Test', 'gross_pay': 2000, 'ytd_cpp': 0, 'ytd_ei': 0}
        on_result = PayrollCalculator(pay_periods=12, province='ON').calculate_payroll(emp)
        ab_result = PayrollCalculator(pay_periods=12, province='AB').calculate_payroll(emp)
        assert ab_result['provincial_tax'] <= on_result['provincial_tax']

    def test_all_provinces_produce_valid_results(self):
        """Every province should produce a valid payroll result."""
        emp = {'name': 'Test', 'gross_pay': 5000, 'ytd_cpp': 0, 'ytd_ei': 0}
        for code in PROVINCES:
            calc = PayrollCalculator(pay_periods=12, province=code)
            result = calc.calculate_payroll(emp)
            assert result['net_amount'] > Decimal('0'), f"{code}: net_amount should be positive"
            assert result['net_amount'] < result['gross_pay'], f"{code}: net should be less than gross"
            assert result['provincial_tax'] >= Decimal('0'), f"{code}: provincial_tax should be non-negative"
