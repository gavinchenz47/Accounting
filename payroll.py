"""CRA 2026 payroll deduction calculator.

Implements the T4127 Payroll Deductions Formulas (122nd Edition)
for CPP, EI, federal tax, and Ontario provincial tax.
"""

import io
from decimal import Decimal, ROUND_HALF_UP


class CRA2026:
    CPP_RATE = Decimal('0.0595')
    CPP_EXEMPTION = Decimal('3500.00')
    CPP_MAX_CONTRIBUTION = Decimal('4230.45')

    EI_RATE = Decimal('0.0163')
    EI_MAX_PREMIUM = Decimal('1123.07')
    EI_EMPLOYER_MULTIPLIER = Decimal('1.4')

    FEDERAL_BPA = Decimal('16452.00')
    ONTARIO_BPA = Decimal('12989.00')

    FEDERAL_BRACKETS = [
        (Decimal('58523'), Decimal('0.14'), Decimal('0')),
        (Decimal('117045'), Decimal('0.205'), Decimal('5826.03')),
        (Decimal('181440'), Decimal('0.26'), Decimal('18606.05')),
        (Decimal('258482'), Decimal('0.29'), Decimal('42030.46')),
        (Decimal('999999999'), Decimal('0.33'), Decimal('61367.74'))
    ]

    ONTARIO_BRACKETS = [
        (Decimal('53359'), Decimal('0.0505'), Decimal('0')),
        (Decimal('106717'), Decimal('0.0915'), Decimal('2187.67')),
        (Decimal('150000'), Decimal('0.1116'), Decimal('7063.32')),
        (Decimal('220000'), Decimal('0.1216'), Decimal('11893.30')),
        (Decimal('999999999'), Decimal('0.1316'), Decimal('20393.30'))
    ]


class PayrollCalculator:
    def __init__(self, pay_periods=12):
        self.P = Decimal(str(pay_periods))
        self.cra = CRA2026()

    def _round(self, amount):
        return Decimal(str(amount)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    def calculate_cpp(self, gross, ytd_cpp):
        exemption = self.cra.CPP_EXEMPTION / self.P
        max_remaining = self.cra.CPP_MAX_CONTRIBUTION - Decimal(str(ytd_cpp))
        period_cpp = self.cra.CPP_RATE * (Decimal(str(gross)) - exemption)
        cpp = min(max_remaining, period_cpp)
        return max(Decimal('0'), self._round(cpp))

    def calculate_ei(self, gross, ytd_ei):
        max_remaining = self.cra.EI_MAX_PREMIUM - Decimal(str(ytd_ei))
        period_ei = self.cra.EI_RATE * Decimal(str(gross))
        ei = min(max_remaining, period_ei)
        return max(Decimal('0'), self._round(ei))

    def calculate_taxes(self, gross, cpp, ei):
        annual_income = Decimal(str(gross)) * self.P

        for bracket_max, rate, constant in self.cra.FEDERAL_BRACKETS:
            if annual_income <= bracket_max:
                R, K = rate, constant
                break

        K1 = Decimal('0.14') * self.cra.FEDERAL_BPA
        cpp_credit = self.P * cpp * (Decimal('0.0495') / Decimal('0.0595'))
        ei_credit = self.P * ei
        K2 = Decimal('0.14') * (cpp_credit + ei_credit)
        K4 = min(Decimal('0.14') * annual_income, Decimal('0.14') * Decimal('1501.00'))

        T3 = max(Decimal('0'), (R * annual_income) - K - K1 - K2 - K4)
        federal_tax = self._round(T3 / self.P)

        for bracket_max, rate, constant in self.cra.ONTARIO_BRACKETS:
            if annual_income <= bracket_max:
                V, KP = rate, constant
                break

        K1P = Decimal('0.0505') * self.cra.ONTARIO_BPA
        K2P = Decimal('0.0505') * (cpp_credit + ei_credit)

        T4 = max(Decimal('0'), (V * annual_income) - KP - K1P - K2P)
        ohp = self._calculate_ohp(annual_income)
        surtax = self._calculate_surtax(T4)

        provincial_tax = self._round((T4 + surtax + ohp) / self.P)

        return federal_tax, provincial_tax

    def _calculate_ohp(self, annual_income):
        A = annual_income
        if A <= Decimal('20000'):
            return Decimal('0')
        elif A <= Decimal('36000'):
            return min(Decimal('300'), Decimal('0.06') * (A - Decimal('20000')))
        elif A <= Decimal('48000'):
            return min(Decimal('450'), Decimal('300') + Decimal('0.06') * (A - Decimal('36000')))
        elif A <= Decimal('72000'):
            return min(Decimal('600'), Decimal('450') + Decimal('0.25') * (A - Decimal('48000')))
        elif A <= Decimal('200000'):
            return min(Decimal('750'), Decimal('600') + Decimal('0.25') * (A - Decimal('72000')))
        else:
            return min(Decimal('900'), Decimal('750') + Decimal('0.25') * (A - Decimal('200000')))

    def _calculate_surtax(self, T4):
        if T4 <= Decimal('5818'):
            return Decimal('0')
        elif T4 <= Decimal('7446'):
            return Decimal('0.20') * (T4 - Decimal('5818'))
        else:
            return Decimal('0.20') * (T4 - Decimal('5818')) + Decimal('0.36') * (T4 - Decimal('7446'))

    def calculate_payroll(self, employee):
        gross = Decimal(str(employee['gross_pay']))
        ytd_cpp = Decimal(str(employee.get('ytd_cpp', 0)))
        ytd_ei = Decimal(str(employee.get('ytd_ei', 0)))

        cpp = self.calculate_cpp(gross, ytd_cpp)
        ei = self.calculate_ei(gross, ytd_ei)
        federal_tax, provincial_tax = self.calculate_taxes(gross, cpp, ei)

        total_deductions = cpp + ei + federal_tax + provincial_tax
        net_amount = gross - total_deductions

        employer_cpp = cpp
        employer_ei = self._round(ei * self.cra.EI_EMPLOYER_MULTIPLIER)
        employer_total = employer_cpp + employer_ei
        remittance_total = total_deductions + employer_total

        return {
            'name': employee['name'],
            'gross_pay': self._round(gross),
            'cpp': cpp,
            'ei': ei,
            'federal_tax': federal_tax,
            'provincial_tax': provincial_tax,
            'total_deductions': total_deductions,
            'net_amount': net_amount,
            'employer_cpp': employer_cpp,
            'employer_ei': employer_ei,
            'employer_total': employer_total,
            'remittance_total': remittance_total,
            'ytd_cpp_new': self._round(ytd_cpp + cpp),
            'ytd_ei_new': self._round(ytd_ei + ei)
        }


def create_excel_output(payroll_data, pay_period):
    """Create Excel file in memory with Payroll Summary and T4 Preparation sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Summary"

    # Title
    ws['A1'] = 'PAYROLL DEDUCTIONS SUMMARY'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws.merge_cells('A1:L1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['A2'] = f'Pay Period: {pay_period}'
    ws['A2'].font = Font(italic=True, size=11)
    ws.merge_cells('A2:L2')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Employee', 'Gross Pay', 'CPP (Employee)', 'EI (Employee)',
               'Federal Tax', 'Provincial Tax', 'Total Deductions', 'Net Amount',
               'CPP (Employer)', 'EI (Employer)', 'Employer Total', 'Remit to CRA']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data
    row = 5
    for emp in payroll_data:
        ws.cell(row=row, column=1, value=emp['name'])
        ws.cell(row=row, column=2, value=float(emp['gross_pay'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=3, value=float(emp['cpp'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=4, value=float(emp['ei'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=5, value=float(emp['federal_tax'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=6, value=float(emp['provincial_tax'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=7, value=float(emp['total_deductions'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=8, value=float(emp['net_amount'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=9, value=float(emp['employer_cpp'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=10, value=float(emp['employer_ei'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=11, value=float(emp['employer_total'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=12, value=float(emp['remittance_total'])).number_format = '$#,##0.00'

        ws.cell(row=row, column=8).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        ws.cell(row=row, column=8).font = Font(bold=True, color='375623')
        row += 1

    # Total row
    total_row = row
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True, size=11)
    for col in range(2, 13):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f'=SUM({chr(64+col)}5:{chr(64+col)}{total_row-1})'
        cell.font = Font(bold=True, size=11)
        cell.number_format = '$#,##0.00'
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # Column widths
    ws.column_dimensions['A'].width = 20
    for col in 'BCDEFGHIJKL':
        ws.column_dimensions[col].width = 14

    # T4 Sheet
    ws_t4 = wb.create_sheet(title="T4 Preparation")
    ws_t4['A1'] = 'T4 SLIP PREPARATION DATA'
    ws_t4['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws_t4['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws_t4.merge_cells('A1:G1')
    ws_t4['A1'].alignment = Alignment(horizontal='center')

    t4_headers = ['Employee Name', 'Box 14\nEmployment Income', 'Box 16\nCPP',
                  'Box 18\nEI', 'Box 22\nIncome Tax', 'YTD CPP', 'YTD EI']

    for col, header in enumerate(t4_headers, 1):
        cell = ws_t4.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    row = 4
    for emp in payroll_data:
        ws_t4.cell(row=row, column=1, value=emp['name'])
        ws_t4.cell(row=row, column=2, value=float(emp['gross_pay'])).number_format = '$#,##0.00'
        ws_t4.cell(row=row, column=3, value=float(emp['cpp'])).number_format = '$#,##0.00'
        ws_t4.cell(row=row, column=4, value=float(emp['ei'])).number_format = '$#,##0.00'
        ws_t4.cell(row=row, column=5, value=float(emp['federal_tax'] + emp['provincial_tax'])).number_format = '$#,##0.00'
        ws_t4.cell(row=row, column=6, value=float(emp['ytd_cpp_new'])).number_format = '$#,##0.00'
        ws_t4.cell(row=row, column=7, value=float(emp['ytd_ei_new'])).number_format = '$#,##0.00'
        row += 1

    ws_t4.column_dimensions['A'].width = 20
    for col in 'BCDEFG':
        ws_t4.column_dimensions[col].width = 16

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
